#!/usr/bin/env python3
"""
Generate Test Evaluation Report with Cosine Similarity

This script processes test inference results and generates an Excel report
with cosine similarity scores between predictions and ground truth.

Usage:
    python scripts/generate_test_report.py --predictions test_predictions.json
    python scripts/generate_test_report.py --predictions test_predictions.json --output test_report.xlsx
"""

import json
import argparse
import numpy as np
from pathlib import Path
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics.pairwise import cosine_similarity
import evaluate
from sentence_transformers import SentenceTransformer


def compute_meteor_score(reference: str, hypothesis: str, metric) -> float:
    """Compute METEOR score."""
    if not reference or not hypothesis or not metric:
        return 0.0

    try:
        return metric.compute(predictions=[hypothesis], references=[reference])['meteor']
    except:
        return 0.0


def compute_cosine_similarity_bert(text1: str, text2: str, model) -> float:
    """Compute cosine similarity using BERT embeddings."""
    if not text1 or not text2:
        return 0.0

    try:
        embeddings = model.encode([text1, text2])
        similarity = cosine_similarity([embeddings[0]], [embeddings[1]])[0][0]
        return float(similarity)
    except:
        return 0.0


def create_excel_report(results: List[Dict], output_path: str, use_bert: bool = True):
    """Create an Excel report with formatted results and similarity scores."""

    # Load BERT model if requested
    bert_model = None
    if use_bert:
        print("Loading BERT model for semantic similarity...")
        try:
            bert_model = SentenceTransformer('all-MiniLM-L6-v2')
            print("✓ BERT model loaded")
        except Exception as e:
            print(f"⚠ Failed to load BERT model: {e}")
            print("  Falling back to TF-IDF similarity")
            use_bert = False

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Evaluation Results"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "ID",
        "Video Path",
        "Ground Truth",
        "Model Prediction",
    ]

    if use_bert:
        headers.append("BERT Similarity")

    headers.append("METEOR Score")
    headers.extend(["Status", "Error"])

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 40  # Video Path
    ws.column_dimensions['C'].width = 50  # Ground Truth
    ws.column_dimensions['D'].width = 50  # Prediction

    if use_bert:
        ws.column_dimensions['E'].width = 18  # BERT Similarity
        ws.column_dimensions['F'].width = 18  # METEOR Score
        ws.column_dimensions['G'].width = 12  # Status
        ws.column_dimensions['H'].width = 40  # Error
    else:
        ws.column_dimensions['E'].width = 18  # METEOR Score
        ws.column_dimensions['F'].width = 12  # Status
        ws.column_dimensions['G'].width = 40  # Error

    # Freeze header row
    ws.freeze_panes = "A2"

    # Process results
    print(f"\nProcessing {len(results)} results...")

    # Load METEOR metric
    meteor_metric = None
    try:
        meteor_metric = evaluate.load('meteor')
        print("✓ METEOR metric loaded")
    except Exception as e:
        print(f"⚠ Failed to load METEOR metric: {e}")

    bert_scores = []
    meteor_scores = []

    for idx, result in enumerate(results, start=1):
        row = idx + 1

        video_path = result.get('video_path', '')
        ground_truth = result.get('ground_truth', '')
        prediction = result.get('prediction', '')
        status = result.get('status', 'unknown')
        error = result.get('error', '')

        # Compute similarities
        bert_sim = None
        if use_bert and bert_model:
            bert_sim = compute_cosine_similarity_bert(ground_truth, prediction, bert_model)
            bert_scores.append(bert_sim)

        meteor_sim = compute_meteor_score(ground_truth, prediction, meteor_metric)
        meteor_scores.append(meteor_sim)

        # Write data
        col = 1
        ws.cell(row=row, column=col).value = idx
        col += 1
        ws.cell(row=row, column=col).value = video_path
        col += 1
        ws.cell(row=row, column=col).value = ground_truth
        col += 1
        ws.cell(row=row, column=col).value = prediction
        col += 1

        if use_bert and bert_sim is not None:
            ws.cell(row=row, column=col).value = round(bert_sim, 4)
            col += 1

        ws.cell(row=row, column=col).value = round(meteor_sim, 4)
        col += 1

        ws.cell(row=row, column=col).value = status
        col += 1
        ws.cell(row=row, column=col).value = error

        # Apply formatting
        for c in range(1, col + 1):
            cell = ws.cell(row=row, column=c)
            cell.alignment = cell_alignment
            cell.border = border

            # Color code similarity scores
            # Determine column indices for coloring
            bert_col = 5 if use_bert else -1
            meteor_col = 6 if use_bert else 5

            if use_bert and c == bert_col:  # BERT column
                score = bert_sim if bert_sim is not None else 0
                if score >= 0.7:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= 0.4:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            if c == meteor_col:  # METEOR column
                score = meteor_sim
                if score >= 0.5: # METEOR scores are often lower than cosine sim
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= 0.2:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")

    summary_data = [
        ["Metric", "Value"],
        ["Total Samples", len(results)],
        ["Successful", sum(1 for r in results if r.get('status') == 'success')],
        ["Failed", sum(1 for r in results if r.get('status') == 'error')],
        ["", ""],
    ]

    if use_bert and bert_scores:
        summary_data.extend([
            ["BERT Similarity", ""],
            ["Mean", round(np.mean(bert_scores), 4)],
            ["Median", round(np.median(bert_scores), 4)],
            ["Std Dev", round(np.std(bert_scores), 4)],
            ["Min", round(np.min(bert_scores), 4)],
            ["Max", round(np.max(bert_scores), 4)],
            ["", ""],
        ])

    if meteor_scores:
        summary_data.extend([
            ["METEOR Score", ""],
            ["Mean", round(np.mean(meteor_scores), 4)],
            ["Median", round(np.median(meteor_scores), 4)],
            ["Std Dev", round(np.std(meteor_scores), 4)],
            ["Min", round(np.min(meteor_scores), 4)],
            ["Max", round(np.max(meteor_scores), 4)],
        ])

    for row_idx, row_data in enumerate(summary_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border

            if row_idx == 1 or (isinstance(value, str) and value and row_idx > 1 and col_idx == 1):
                cell.font = Font(bold=True)

    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 15

    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel report saved to: {output_path}")

    # Print summary
    print(f"\n{'='*60}")
    print("Evaluation Summary")
    print(f"{'='*60}")
    print(f"Total samples: {len(results)}")
    print(f"Successful: {sum(1 for r in results if r.get('status') == 'success')}")
    print(f"Failed: {sum(1 for r in results if r.get('status') == 'error')}")

    if use_bert and bert_scores:
        print(f"\nBERT Similarity:")
        print(f"  Mean: {np.mean(bert_scores):.4f}")
        print(f"  Median: {np.median(bert_scores):.4f}")
        print(f"  Std Dev: {np.std(bert_scores):.4f}")

    if meteor_scores:
        print(f"\nMETEOR Score:")
        print(f"  Mean: {np.mean(meteor_scores):.4f}")
        print(f"  Median: {np.median(meteor_scores):.4f}")
        print(f"  Std Dev: {np.std(meteor_scores):.4f}")

    print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser(description="Generate test evaluation report with similarity scores")
    parser.add_argument("--predictions", type=str, required=True,
                        help="Path to predictions JSON from test_inference.py")
    parser.add_argument("--output", type=str, default=None,
                        help="Output Excel file path (default: same directory as predictions)")
    parser.add_argument("--no-bert", action="store_true",
                        help="Skip BERT similarity (faster, uses only TF-IDF)")

    args = parser.parse_args()

    # Set default output path to same directory as predictions if not provided
    if args.output is None:
        pred_path = Path(args.predictions)
        args.output = str(pred_path.parent / "test_evaluation_report.xlsx")
        print(f"Output will be saved to: {args.output}")

    # Load predictions
    print(f"Loading predictions from: {args.predictions}")
    with open(args.predictions, 'r') as f:
        results = json.load(f)

    print(f"Loaded {len(results)} predictions")

    # Generate report
    create_excel_report(results, args.output, use_bert=not args.no_bert)


if __name__ == "__main__":
    main()
